"""
Convert Custom_Cabinetry_Products.xlsx into a normalized JSON seed.

Two passes:
  1. classify every attribute column by observing all its values across all sheets
     into one of {numeric, bool, brand, enum, text}
  2. emit the seed JSON, routing each attribute value to the matching typed slot

The image attribute ("Основное изображение") is treated specially: it never reaches
`products[].attributes[]` — instead the URL is hashed to a stable filename and stored
as `products[].mainImageFilename`. Images are downloaded separately by
`download_images.py` into wwwroot/.

Output schema (matches the EF Core model in ProductsService):
{
  "categories":  [{id, parentId, code, titleRu, isLeaf}],
  "brands":      [{id, titleRu}],
  "suppliers":   [{id, name}],
  "attributes":  [{id, code, titleRu, valueType}],    # excludes the image attr
  "enumValues":  [{id, attributeId, code, titleRu, sortOrder}],
  "products":    [{id, externalCode, nameRu, categoryId, mainImageFilename?,
                    attributes: [{attrId, valueText?, valueNumeric?, valueBool?,
                                  enumValueId?, brandId?, valueType}],
                    offers:     [{supplierId, priceAmount, currency}]}]
}

Run modes:
  python xlsx_to_seed.py            # full emit → writes seed.json
  python xlsx_to_seed.py --classify # dry-run, prints classifier table, no file I/O
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

import openpyxl

XLSX = Path(r"C:/Users/kalmy/Downloads/Custom_Cabinetry_Products.xlsx")
OUT = Path(
    r"C:/Users/kalmy/OneDrive/Desktop/domeoProductsDb/DomeoProductsDb/"
    r"src/DomeoProductsDb.ProductsService/SeedData/seed.json"
)
OVERRIDES = Path(__file__).with_name("attribute_type_overrides.json")

NON_ATTRIBUTE_HEADERS = {"Артикул", "Наименование", "Цена ₽", "Поставщик"}

# Attribute column holding the image URL. It is lifted out of attributes[] into
# the top-level Product.mainImageFilename (hashed). See download_images.py for
# the matching downloader that saves files under wwwroot/images/{full,preview}/.
IMAGE_HEADER = "Основное изображение"


def url_to_filename(url: str) -> str:
    """Stable 16-hex-char sha1 prefix + .jpg. Safe for 302 unique URLs."""
    digest = hashlib.sha1(url.strip().encode("utf-8")).hexdigest()[:16]
    return f"{digest}.jpg"

BOOL_TRUE = {"да", "yes", "true"}
BOOL_FALSE = {"нет", "no", "false"}

ENUM_MAX_DISTINCT = 30
ENUM_MAX_LEN = 60


def slugify(title: str) -> str:
    t = title.strip().lower()
    t = re.sub(r"[^\w]+", "_", t, flags=re.UNICODE)
    return t.strip("_") or "x"


def load_overrides() -> dict[str, str]:
    if OVERRIDES.exists():
        return json.loads(OVERRIDES.read_text(encoding="utf-8"))
    return {}


def is_numeric(v) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", "."))
        return True
    except (TypeError, ValueError):
        return False


def is_bool_literal(v) -> bool:
    if isinstance(v, bool):
        return True
    if v is None:
        return False
    return str(v).strip().lower() in BOOL_TRUE | BOOL_FALSE


def classify_column(header: str, values: list) -> str:
    """Return one of: numeric, bool, brand, enum, text."""
    if header == "Бренд":
        return "brand"
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "text"
    if all(is_bool_literal(v) for v in non_null):
        return "bool"
    if all(is_numeric(v) for v in non_null):
        return "numeric"
    str_vals = [str(v).strip() for v in non_null]
    distinct = set(str_vals)
    if len(distinct) <= ENUM_MAX_DISTINCT and all(len(s) <= ENUM_MAX_LEN for s in distinct):
        return "enum"
    return "text"


def iter_product_sheets(wb):
    for ws in wb.worksheets:
        if ws.title == "Сводка":
            continue
        yield ws


def collect_attribute_values(wb) -> dict[str, list]:
    """Return {header_title: [all observed non-empty values across all sheets]}."""
    by_header: dict[str, list] = {}
    for ws in iter_product_sheets(wb):
        headers = [c.value for c in ws[1]]
        for idx, header in enumerate(headers):
            if not header or header in NON_ATTRIBUTE_HEADERS:
                continue
            bucket = by_header.setdefault(header, [])
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                v = row[idx]
                if v is not None and v != "":
                    bucket.append(v)
    return by_header


def run_classify(wb) -> None:
    overrides = load_overrides()
    cols = collect_attribute_values(wb)
    counts = {"numeric": 0, "bool": 0, "brand": 0, "enum": 0, "text": 0}
    print(f"{'header':<45} {'type':<8} {'dist':>5} {'maxlen':>6}  samples")
    print("-" * 120)
    for header in sorted(cols):
        values = cols[header]
        inferred = classify_column(header, values)
        code = slugify(header)
        final = overrides.get(code, inferred)
        counts[final] = counts.get(final, 0) + 1
        str_vals = [str(v) for v in values]
        distinct = sorted(set(str_vals))
        max_len = max((len(s) for s in distinct), default=0)
        tag = final if final == inferred else f"{final}*"
        sample = ", ".join(distinct[:5]) + (" …" if len(distinct) > 5 else "")
        print(f"{header[:45]:<45} {tag:<8} {len(distinct):>5} {max_len:>6}  {sample[:55]}")
    print("-" * 120)
    print("counts:", counts)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--classify", action="store_true", help="dry-run: print classifier table")
    parser.add_argument("out", nargs="?", type=Path, default=OUT)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    if args.classify:
        run_classify(wb)
        return

    overrides = load_overrides()

    # ---- Pass 1: classify every attribute column ----
    cols = collect_attribute_values(wb)
    attr_types: dict[str, str] = {}
    for header, values in cols.items():
        inferred = classify_column(header, values)
        attr_types[header] = overrides.get(slugify(header), inferred)

    # ---- Categories from Сводка ----
    categories: list[dict] = []
    summary = wb["Сводка"]
    for row in summary.iter_rows(min_row=2, values_only=True):
        cid, parent_id, code, title, ctype, is_leaf_ru, *_ = row
        if not isinstance(cid, int):
            continue
        categories.append({
            "id": int(cid),
            "parentId": int(parent_id) if parent_id is not None else None,
            "code": code,
            "titleRu": title,
            "isLeaf": (is_leaf_ru == "Да"),
        })

    leaf_cats = [c for c in categories if c["isLeaf"]]
    title_to_cat = {c["titleRu"]: c for c in leaf_cats}

    def find_category_for_sheet(sheet_title: str) -> dict | None:
        if sheet_title in title_to_cat:
            return title_to_cat[sheet_title]
        stem = sheet_title.rstrip(".").rstrip()
        for cat_title, cat in title_to_cat.items():
            clean = cat_title.replace("/", "")
            if clean == stem or clean.startswith(stem) or stem.startswith(clean[:28]):
                return cat
        return None

    # ---- registries ----
    suppliers: dict[str, int] = {}
    brands: dict[str, int] = {}
    attributes: dict[str, dict] = {}  # code -> {id, code, titleRu, valueType}
    enum_values: dict[tuple[int, str], dict] = {}  # (attrId, valueTitle) -> {id, ...}
    products: list[dict] = []

    next_supplier_id = 1
    next_brand_id = 1
    next_attr_id = 1
    next_enum_id = 1
    next_product_id = 1

    def supplier_id(name: str) -> int | None:
        nonlocal next_supplier_id
        if not name:
            return None
        if name not in suppliers:
            suppliers[name] = next_supplier_id
            next_supplier_id += 1
        return suppliers[name]

    def brand_id(title: str) -> int | None:
        nonlocal next_brand_id
        if not title:
            return None
        if title not in brands:
            brands[title] = next_brand_id
            next_brand_id += 1
        return brands[title]

    def ensure_attr(header: str) -> dict:
        nonlocal next_attr_id
        code = slugify(header)
        if code not in attributes:
            value_type = attr_types.get(header, "text")
            attributes[code] = {
                "id": next_attr_id,
                "code": code,
                "titleRu": header,
                "valueType": value_type,
            }
            next_attr_id += 1
            if value_type == "enum":
                # pre-register all enum values in sorted order so SortOrder is stable
                distinct_titles = sorted(
                    {str(v).strip() for v in cols.get(header, []) if v is not None and v != ""}
                )
                for order, title in enumerate(distinct_titles):
                    register_enum_value(attributes[code]["id"], title, order)
        return attributes[code]

    def register_enum_value(attr_id: int, title: str, sort_order: int) -> int:
        nonlocal next_enum_id
        key = (attr_id, title)
        if key not in enum_values:
            enum_values[key] = {
                "id": next_enum_id,
                "attributeId": attr_id,
                "code": slugify(title),
                "titleRu": title,
                "sortOrder": sort_order,
            }
            next_enum_id += 1
        return enum_values[key]["id"]

    for ws in iter_product_sheets(wb):
        cat = find_category_for_sheet(ws.title)
        if cat is None:
            print(f"WARN: no category match for sheet {ws.title!r}")
            continue

        headers = [c.value for c in ws[1]]
        for h in headers:
            if h and h not in NON_ATTRIBUTE_HEADERS and h != IMAGE_HEADER:
                ensure_attr(h)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            product = {
                "id": next_product_id,
                "externalCode": str(row[0]),
                "nameRu": str(row[1]) if row[1] is not None else "",
                "categoryId": cat["id"],
                "mainImageFilename": None,
                "attributes": [],
                "offers": [],
            }
            next_product_id += 1

            price = row[2]
            supplier = row[3]
            if price is not None and supplier:
                product["offers"].append({
                    "supplierId": supplier_id(str(supplier)),
                    "priceAmount": float(price),
                    "currency": "RUB",
                })

            for idx, header in enumerate(headers):
                if not header or header in NON_ATTRIBUTE_HEADERS:
                    continue
                value = row[idx]
                if value is None or value == "":
                    continue
                # Image attribute: lift out of attributes[], hash URL → filename.
                if header == IMAGE_HEADER:
                    product["mainImageFilename"] = url_to_filename(str(value))
                    continue
                attr = ensure_attr(header)
                av: dict = {"attrId": attr["id"], "valueType": attr["valueType"]}
                t = attr["valueType"]
                if t == "brand":
                    av["brandId"] = brand_id(str(value))
                elif t == "bool":
                    if isinstance(value, bool):
                        av["valueBool"] = value
                    else:
                        sv = str(value).strip().lower()
                        av["valueBool"] = sv in BOOL_TRUE
                elif t == "numeric":
                    av["valueNumeric"] = float(str(value).replace(",", "."))
                elif t == "enum":
                    title = str(value).strip()
                    ev_id = register_enum_value(attr["id"], title, 0)  # order set during ensure_attr
                    av["enumValueId"] = ev_id
                else:  # text
                    av["valueText"] = str(value).strip()
                product["attributes"].append(av)

            products.append(product)

    seed = {
        "categories": categories,
        "brands": [{"id": i, "titleRu": t} for t, i in brands.items()],
        "suppliers": [{"id": i, "name": n} for n, i in suppliers.items()],
        "attributes": list(attributes.values()),
        "enumValues": list(enum_values.values()),
        "products": products,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(seed, ensure_ascii=False), encoding="utf-8")
    print(
        f"Wrote {args.out}: "
        f"{len(categories)} categories, {len(brands)} brands, "
        f"{len(suppliers)} suppliers, {len(attributes)} attributes, "
        f"{len(enum_values)} enum values, {len(products)} products"
    )


if __name__ == "__main__":
    main()
