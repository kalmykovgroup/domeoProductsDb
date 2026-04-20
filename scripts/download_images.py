"""
Download every unique product image URL from seed.json into wwwroot/images/full/
and generate a 300×300 preview under wwwroot/images/preview/.

Idempotent — already-downloaded/resized files are skipped.
Produces a sidecar `scripts/image_status.json` with any URLs that failed.

Filenames are sha1(url)[:16] + '.jpg' — matches the scheme used by
xlsx_to_seed.py::url_to_filename() and Product.MainImageFilename in the DB.
"""
from __future__ import annotations

import concurrent.futures
import hashlib
import io
import json
import time
from pathlib import Path

import requests
from PIL import Image

SEED = Path(
    r"C:/Users/kalmy/OneDrive/Desktop/domeoProductsDb/DomeoProductsDb/"
    r"src/DomeoProductsDb.ProductsService/SeedData/seed.json"
)
WWWROOT = Path(
    r"C:/Users/kalmy/OneDrive/Desktop/domeoProductsDb/DomeoProductsDb/"
    r"src/DomeoProductsDb.ProductsService/wwwroot"
)
STATUS_FILE = Path(__file__).with_name("image_status.json")

FULL_DIR = WWWROOT / "images" / "full"
PREVIEW_DIR = WWWROOT / "images" / "preview"
PREVIEW_SIZE = (300, 300)
REQUEST_TIMEOUT = 20
MAX_WORKERS = 8
MAX_RETRIES = 2


def url_to_filename(url: str) -> str:
    digest = hashlib.sha1(url.strip().encode("utf-8")).hexdigest()[:16]
    return f"{digest}.jpg"


def collect_urls() -> dict[str, str]:
    """Return {url: filename} from the raw XLSX — we need the URL, not just the
    filename stored in seed.json. Re-parse via the mapping: same sha1 function."""
    # We don't have the original URLs in seed.json anymore (dropped from
    # attributes[]). Re-read the XLSX directly — it's the source of truth.
    import openpyxl
    xlsx = Path(r"C:/Users/kalmy/Downloads/Custom_Cabinetry_Products.xlsx")
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    urls: dict[str, str] = {}
    image_col_name = "Основное изображение"
    for ws in wb.worksheets:
        if ws.title == "Сводка":
            continue
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        try:
            idx = headers.index(image_col_name)
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[idx] if idx < len(row) else None
            if v:
                u = str(v).strip()
                urls[u] = url_to_filename(u)
    return urls


def download_one(url: str, filename: str) -> tuple[str, str | None]:
    """Download URL to full/, then resize to preview/. Returns (status, error).
    status ∈ {'downloaded', 'skipped', 'failed'}."""
    full = FULL_DIR / filename
    preview = PREVIEW_DIR / filename
    if full.exists() and preview.exists():
        return "skipped", None

    for attempt in range(MAX_RETRIES + 1):
        try:
            if not full.exists():
                r = requests.get(
                    url,
                    timeout=REQUEST_TIMEOUT,
                    headers={"User-Agent": "DomeoProductsDb-Seeder/1.0"},
                )
                r.raise_for_status()
                if not r.headers.get("Content-Type", "").startswith("image/"):
                    return "failed", f"not image ({r.headers.get('Content-Type')})"
                full.write_bytes(r.content)

            if not preview.exists():
                with Image.open(full) as im:
                    im = im.convert("RGB")
                    im.thumbnail(PREVIEW_SIZE, Image.LANCZOS)
                    im.save(preview, format="JPEG", quality=85, optimize=True)

            return "downloaded", None
        except requests.RequestException as e:
            if attempt < MAX_RETRIES:
                time.sleep(1 + attempt)
                continue
            return "failed", f"http: {e}"
        except Exception as e:
            return "failed", f"pillow: {e}"
    return "failed", "exhausted retries"


def main() -> None:
    FULL_DIR.mkdir(parents=True, exist_ok=True)
    PREVIEW_DIR.mkdir(parents=True, exist_ok=True)

    urls = collect_urls()
    print(f"Found {len(urls)} unique image URLs")

    counts = {"downloaded": 0, "skipped": 0, "failed": 0}
    failures: dict[str, str] = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = {pool.submit(download_one, u, fn): (u, fn) for u, fn in urls.items()}
        for i, fut in enumerate(concurrent.futures.as_completed(futures), start=1):
            url, fn = futures[fut]
            status, err = fut.result()
            counts[status] += 1
            if status == "failed":
                failures[url] = err or "unknown"
                print(f"  [{i:>3}/{len(urls)}] FAIL {url} — {err}")
            else:
                print(f"  [{i:>3}/{len(urls)}] {status} {fn}")

    print()
    print(f"Summary: downloaded={counts['downloaded']}  skipped={counts['skipped']}  failed={counts['failed']}")
    STATUS_FILE.write_text(json.dumps({
        "summary": counts,
        "failed": failures,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Status saved: {STATUS_FILE}")


if __name__ == "__main__":
    main()
